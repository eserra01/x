# -*- encoding: utf-8 -*-
##################################################################################################
#
#   Author: Experts SRL de CV (https://exdoo.mx)
#   Coded by: Rodolfo Lopez (rodolfo.lopez@exdoo.mx)
#   Migrated by: Daniel Acosta (daniel.acosta@exdoo.mx)
#   License: https://blog.exdoo.mx/licencia-de-uso-de-software/
#
##################################################################################################
from odoo import models, fields,_
import openpyxl
import base64
from io import BytesIO
from odoo.exceptions import UserError
import logging
# pip install openpyxl

_logger = logging.getLogger(__name__)

class ImportXLSWizard(models.Model):
    _name = 'import.xls.wizard'
    
    file = fields.Binary(string="Archivo", required=True)
    file_name = fields.Char(string="File Name")     
    info = fields.Text(string="Info", readonly=True, default= '...')   
    data = fields.Text(string="Data")   
    company_id = fields.Many2one('res.company', 'Company', required=True, index=True, default=lambda self: self.env.company)
    option = fields.Selection([('bt','Bitácora'),('ct','Árbol de comisiones'),('cm','Comisiones')],string="Aplicación")

    def import_file(self):       
        wb = openpyxl.load_workbook( 
        filename=BytesIO(base64.b64decode(self.file)), read_only=True)
        ws = wb.active
        if self.option == 'bt':
            vals = []
            info = ''
            for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=None, values_only=True):                                    
                # Se busca la bitácora 
                mortuary_id = self.env['mortuary'].search([('name','=',record[0])])
                if not mortuary_id:
                    # Hora de llamada
                    if len(str(record[4])) == 3:
                        ii_hora_creacion = '0' + str(record[4])
                        ii_hora_creacion = ii_hora_creacion[0:2] + ':' + ii_hora_creacion[2:4]
                    elif len(str(record[4])) == 4 and record[4] not in ['NULL',None]:
                        ii_hora_creacion = str(record[4])[0:2] + ':' + str(record[4])[2:4]
                    else:
                        ii_hora_creacion = ''
                    # ii_llamada
                    ii_llamada = False
                    if record[5] == 'SI':
                        ii_llamada = 1
                    if record[5] == 'NO':
                        ii_llamada = 2                
                    # Causa fallecimiento         
                    ii_causa_fallecim_id = False      
                    ii_causa_fallecim = self.env['ii.causa.fallecim'].search([('name','=ilike', str(record[7]))])
                    if ii_causa_fallecim:
                        ii_causa_fallecim_id = ii_causa_fallecim.id
                    # Atiende servicio
                    ds_atiende_servicio_id = False      
                    ds_atiende_servicio = self.env['ds.atiende.servicio'].search([('name','=', str(record[8]))])
                    if ds_atiende_servicio:
                        ds_atiende_servicio_id = ds_atiende_servicio.id
                    else:
                        if str(record[8]) not in ['NULL','- -','None']:                
                            ds_atiende_servicio = self.env['ds.atiende.servicio'].create({'name': str(record[8]).upper()})                                     
                            ds_atiende_servicio_id = ds_atiende_servicio.id 
                        else:
                            ds_atiende_servicio_id = False
                    # Servicio confirmado
                    cs_servi_confirm_id = False
                    if str(record[11]) == 'SI':
                        cs_servi_confirm_id = 1 
                    if str(record[11]) == 'NO':
                        cs_servi_confirm_id = 2 
                    # Hora confirmación
                    if len(str(record[16])) == 3:
                        cs_hora_confirm = '0' + str(record[16])
                        # cs_hora_confirm = cs_hora_confirm[0:2] + ':' + cs_hora_confirm[2:4]
                    elif len(str(record[16])) == 4 and record[16] not in ['NULL',None]:
                        cs_hora_confirm = str(record[16])[0:2] + str(record[16])[2:4]                         
                    else:
                        cs_hora_confirm = False
                    # Tipo de servicio
                    ds_tipo_de_servicio_id = False
                    ds_tipo_de_servicio = self.env['ds.tipo.servicio'].search([('name','=ilike', str(record[19]))])               
                    if ds_tipo_de_servicio:                   
                        ds_tipo_de_servicio_id = ds_tipo_de_servicio.id
                    else:
                        if str(record[19]) not in ['NULL','- -','None']:
                            ds_tipo_de_servicio = self.env['ds.tipo.servicio'].create({'name': str(record[19]).upper()})    
                            ds_tipo_de_servicio_id = ds_tipo_de_servicio.id
                    # ds_origen
                    ds_origen_id = False
                    ds_origen = self.env['ds.origen'].search([('name','=ilike', str(record[20]))])               
                    if ds_origen:                   
                        ds_origen_id = ds_origen.id                   
                    else:
                        if str(record[20]) not in ['NULL','- -','None']:
                            ds_origen = self.env['ds.origen'].create({'name': str(record[20]).upper()})    
                            ds_origen_id = ds_origen.id     
                    # ds_interplaza
                    ds_interplaza_id = False                           
                    if str(record[21]) in ['- - -']:                   
                        ds_interplaza_id = 2                
                    else:
                        ds_interplaza_id = 1       
                    # Aplica seguro
                    ds_aplica_seguro_id = False
                    if str(record[26]) == 'SI':
                        ds_aplica_seguro_id = 1 
                    if str(record[26]) == 'NO':
                        ds_aplica_seguro_id = 2
                    # Ataud
                    ds_ataud_id = False
                    ds_ataud = self.env['ds.ataud'].search([('name','=ilike', str(record[38]))])
                    if ds_ataud:
                        ds_ataud_id = ds_ataud.id
                    else:
                        if str(record[38]) not in ['NULL','----','None']:
                            ds_ataud = self.env['ds.ataud'].create({'name': str(record[38]).upper()})
                            ds_ataud_id = ds_ataud.id
                    # Urna
                    ds_urna_id = False
                    ds_urna = self.env['ds.urna'].search([('name','=ilike', str(record[39]))])
                    if ds_urna:
                        ds_urna_id = ds_urna.id
                    else:
                        if str(record[39]) not in ['NULL','----','None']:
                            ds_urna = self.env['ds.urna'].create({'name': str(record[39]).upper()})
                            ds_urna_id = ds_urna.id
                    #
                    dc_forma_de_pago_id = False
                    dc_forma_de_pago = self.env['dc.forma.pago'].search([('name','=ilike', str(record[42]))])
                    if dc_forma_de_pago:
                        dc_forma_de_pago_id = dc_forma_de_pago.id
                    else:
                        if str(record[42]) not in ['NULL','----','None']:
                            dc_forma_de_pago = self.env['dc.forma.pago'].create({'name': str(record[42]).upper()})
                            dc_forma_de_pago_id = dc_forma_de_pago.id
                    # lugar de velación
                    iv_lugar_de_velacion_id = False
                    iv_lugar_de_velacion = self.env['iv.lugar.velacion'].search([('name','=ilike', str(record[45]))])
                    if iv_lugar_de_velacion:
                        iv_lugar_de_velacion_id = iv_lugar_de_velacion.id
                    else:
                        if str(record[45]) not in ['NULL','None']:
                            iv_lugar_de_velacion = self.env['iv.lugar.velacion'].create({'name': str(record[45]).upper()})
                            iv_lugar_de_velacion_id = iv_lugar_de_velacion.id
                    # nombre capilla
                    iv_nombre_de_capilla_id = False
                    iv_nombre_de_capilla = self.env['iv.nombre.capilla'].search([('name','=ilike', str(record[46]))])
                    if iv_nombre_de_capilla:
                        iv_nombre_de_capilla_id = iv_nombre_de_capilla.id
                    else:
                        if str(record[46]) not in ['NULL','None']:
                            iv_nombre_de_capilla = self.env['iv.nombre.capilla'].create({'name': str(record[46]).upper()})
                            iv_nombre_de_capilla_id = iv_nombre_de_capilla.id
                    # Ropa
                    ig_entrego_ropa_id = False
                    if str(record[52]) == 'SI':
                        ig_entrego_ropa_id = 1 
                    if str(record[52]) == 'NO':
                        ig_entrego_ropa_id = 2
                    # operativo 1
                    ir_operativo_1_id = False
                    ir_operativo_1 = self.env['ir.operativo'].search([('name','=ilike', str(record[53]))])
                    if ir_operativo_1:
                        ir_operativo_1_id = ir_operativo_1.id
                    else:
                        if str(record[53]) not in ['NULL','None']:
                            ir_operativo_1 = self.env['ir.operativo'].create({'name': str(record[53]).upper()})
                            ir_operativo_1_id = ir_operativo_1.id
                    # operativo 2
                    ir_operativo_2_id = False
                    ir_operativo_2 = self.env['ir.operativo'].search([('name','=ilike', str(record[54]))])
                    if ir_operativo_2:
                        ir_operativo_2_id = ir_operativo_2.id
                    else:
                        if str(record[54]) not in ['NULL','None']:
                            ir_operativo_2 = self.env['ir.operativo'].create({'name': str(record[54]).upper()})
                            ir_operativo_2_id = ir_operativo_2.id
                    # Hora de inciio
                    if len(str(record[55])) == 3:
                        ir_hora_de_inicio = '0' + str(record[55])
                        ir_hora_de_inicio = ir_hora_de_inicio[0:2] + ir_hora_de_inicio[2:4]
                    elif len(str(record[55])) == 4 and record[55] not in ['NULL',None]:
                        ir_hora_de_inicio = str(record[55])[0:2] + str(record[55])[2:4]
                    else:
                        ir_hora_de_inicio = ''
                    # carroza i
                    ii_carroza_id = False
                    ii_carroza = self.env['carroza'].search([('name','=ilike', str(record[60]))])
                    if ii_carroza:
                        ii_carroza_id = ii_carroza.id
                    else:
                        if str(record[60]) not in ['NULL','None']:
                            ii_carroza = self.env['carroza'].create({'name': str(record[60]).upper()})
                            ii_carroza_id = ii_carroza.id
                    # carroza c
                    ic_carroza_id = False
                    ic_carroza = self.env['carroza'].search([('name','=ilike', str(record[61]))])
                    if ic_carroza:
                        ic_carroza_id = ic_carroza.id
                    else:
                        if str(record[61]) not in ['NULL','None']:
                            ic_carroza = self.env['carroza'].create({'name': str(record[61]).upper()})
                            ic_carroza_id = ic_carroza.id
                    # relación finado 1
                    contact_1_relacion_confinad_id = False
                    contact_1_relacion_confinad = self.env['relacion.confinad'].search([('name','=ilike', str(record[64]))])
                    if contact_1_relacion_confinad:
                        contact_1_relacion_confinad_id = contact_1_relacion_confinad.id
                    else:
                        if str(record[64]) not in ['NULL','None']:
                            contact_1_relacion_confinad = self.env['relacion.confinad'].create({'name': str(record[64]).upper()})
                            contact_1_relacion_confinad_id = contact_1_relacion_confinad.id                    
                  
                    #
                    val = {
                        'name': record[0],
                        'ii_finado': record[1],
                        'ii_fecha_creacion': str(record[3])[0:10] if record[3] not in ['NULL',None] else False,
                        'ii_hora_creacion':ii_hora_creacion,
                        'ii_llamada': ii_llamada,
                        'ii_folio_certificad': record[6] if record[6] not in ['NULL',None] else '',
                        'ii_causa_fallecim': ii_causa_fallecim_id,
                        'ds_atiende_servicio': ds_atiende_servicio_id,
                        'ii_lugar_fallec': str(record[9]) if record[9] not in ['NULL',None] else '',
                        'ii_direcc_fallecimiento': str(record[10]) if record[10] not in ['NULL',None] else '',
                        'cs_servi_confirm': cs_servi_confirm_id,
                        'cs_agente_confir': str(record[12]) if record[12] not in ['NULL',None] else '',
                        'cs_cliente_confir': str(record[13]) if record[13] not in ['NULL',None] else '',
                        'cs_fecha_confirm': str(record[15])[0:10] if record[15] not in ['NULL',None] else False,
                        'cs_hora_confirm': cs_hora_confirm,
                        'tc_nomb_titular': str(record[17]) if record[17] not in ['NULL',None] else '',
                        'tc_no_contrato': str(record[18]) if record[18] not in ['NULL',None] else '',
                        'ds_tipo_de_servicio': ds_tipo_de_servicio_id,
                        'ds_origen': ds_origen_id,
                        'ds_interplaza': ds_interplaza_id,
                        'ds_personas_autorizadas': str(record[23]) if record[23] not in ['NULL',None] else '',
                        'psa_servi_adicionals': str(record[24]) if record[24] not in ['NULL',None] else '',
                        'ds_aplica_seguro': ds_aplica_seguro_id,
                        'psa_saldo_PABS': float(record[27]) if record[27] not in ['NULL',None] else 0,
                        'psa_ataud_o_cambio': float(record[28]) if record[28] not in ['NULL',None] else 0,
                        'psa_cremacion': float(record[29]) if record[29] not in ['NULL',None] else 0,
                        'psa_certificado': float(record[30]) if record[30] not in ['NULL',None] else 0,
                        'psa_embalsamado': float(record[31])  if record[31] not in ['NULL',None] else 0,
                        'psa_capilla_domicilio': float(record[32])  if record[32] not in ['NULL',None] else 0,
                        'psa_capilla_recinto': float(record[33]) if record[33] not in ['NULL',None] else 0,
                        'psa_cafeteria': float(record[34])  if record[34] not in ['NULL',None] else 0,
                        'psa_traslado': float(record[35]) if record[35] not in ['NULL',None] else 0,
                        'psa_tramites': float(record[36])  if record[36] not in ['NULL',None] else 0,
                        'psa_camion': float(record[37]) if record[37] not in ['NULL',None] else 0,
                        'ds_ataud': ds_ataud_id,
                        'ds_urna': ds_urna_id,      
                        'dc_saldo_conveniado': float(record[41]) if record[41] not in ['NULL',None] else 0,   
                        'dc_forma_de_pago': dc_forma_de_pago_id,
                        'dc_cantidad_de_pagos': float(record[43]) if record[43] not in ['NULL',None] else 0,
                        'dc_fecha_de_inicio': False,
                        'iv_lugar_de_velacion': iv_lugar_de_velacion_id,
                        'iv_nombre_de_capilla': iv_nombre_de_capilla_id,
                        'iv_direccion': str(record[47]) if record[47] not in ['NULL',None] else '',
                        'ig_panteon': str(record[51]) if record[51] not in ['NULL',None] else '',
                        'ig_entrego_ropa': ig_entrego_ropa_id,
                        'ir_operativo_1': ir_operativo_1_id,
                        'ir_operativo_2': ir_operativo_2_id,
                        'ir_hora_de_inicio': ir_hora_de_inicio,
                        'ii_carroza': ii_carroza_id,
                        'ic_carroza': ic_carroza_id,
                        'contact_1_nomb': str(record[63]) if record[63] not in ['NULL',None] else '',
                        'contact_1_relacion_confinad': contact_1_relacion_confinad_id,
                        'contact_1_tel': str(record[65])[0:10] if record[54]not in ['NULL',None] and len(str(record[65])[0:10])==10 else '',
                        'podp_nomb': str(record[66]) if record[55] != 'NULL' else '',
                        'podp_calle_y_number': str(record[67]) if record[67] not in ['NULL',None] else '',
                        'podp_municipio_id': False, # 69
                        'podp_colonia_id': False, # 70,
                        'podp_tel': str(record[72])[0:10] if record[72] not in ['NULL',None] else '',
                    }
                    vals.append(val)                    
                    print(val)      
                    self.env['mortuary'].create(val)                             
                else:
                    info += '\nYa existe la bitácora número %s'%(str(record[0]))
                # break
            #
            # self.env['mortuary'].create(vals)        
            # info += '\nSe crearon %s registros'%(str(len(vals)))
            self.info = info
            #
            return {
                    'name':"Importar XLS",
                    'view_type': 'form',
                    'view_mode': 'form',
                    'view_id': False,
                    'res_model': 'import.xls.wizard',
                    'domain': [],
                    'type': 'ir.actions.act_window',
                    'target': 'new',
                    'res_id': self._ids[0],
            }
        #
        if self.option == 'ct':            
            info = ''
            contracts_not_found = []
            agents_not_found = []
            
            i = 0
            contract_id = 0
            job_id = 0
            pay_order = 0
            #
            last_contract_id = 0
            last_job_id = 0
            last_pay_order = 0
            for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=None, values_only=True):                        
                # Se busca contrato 
                contract_id = self.env['pabs.contract'].search([('name','ilike',record[3]),('company_id','=',self.company_id.id)])
                if contract_id:
                    #                                 
                    # job_id = self.env['hr.job'].search([('name','=',record[2]),('company_id','=',self.company_id.id)], limit = 1)
                    name = record[5]
                    if name == 'PAPELERIA':
                        name = 'PAPELERIA PABS'
                    if name == 'FIDEICOMISO':
                        name = 'FIDEICOMISO PABS' 
                    agent_id = self.env['hr.employee'].search([('barcode','=',record[4]),('company_id','=',self.company_id.id)], limit = 1)                               
                    #
                    jobs = {
                        'ASISTENTE SOCIAL': 810,
                        'RECOMENDADO': 811,
                        'COORDINADOR': 812,
                        'GERENTE DE OFICINA': 813,
                        'FIDEICOMISO': 814,
                        'PAPELERIA': 815,
                        'COBRADOR': 816,

                    }
                    if agent_id:
                        vals = {
                            'contract_id': contract_id.id,
                            'pay_order': record[1],
                            'job_id': jobs.get(record[2]),
                            'comission_agent_id': agent_id.id,
                            'corresponding_commission': record[6],
                            'remaining_commission': record[7],
                            'commission_paid': record[8],
                            'actual_commission_paid': record[9],
                        }            
                        #
                        if last_contract_id == contract_id and last_pay_order == record[1] and last_job_id == jobs.get(record[2]) and i>0:
                            print("CONTINUE")
                            continue                            
                                                                                     
                        print(vals)                       
                        self.env['pabs.comission.tree'].create(vals)
                        i += 1
                        last_contract_id = contract_id 
                        last_pay_order = record[1]
                        last_job_id = jobs.get(record[2])
                        print("Creando comisión %s "%(i))                                                                          
                    else:                                 
                        if name not in agents_not_found:                                       
                            agents_not_found.append(name)                           
                else:
                    if record[3] not in contracts_not_found:                      
                        contracts_not_found.append(record[3])
            #
            self.info = str(contracts_not_found) + "\n" + str(agents_not_found) 
            #
            return {
                    'name':"Importar XLS",
                    'view_type': 'form',
                    'view_mode': 'form',
                    'view_id': False,
                    'res_model': 'import.xls.wizard',
                    'domain': [],
                    'type': 'ir.actions.act_window',
                    'target': 'new',
                    'res_id': self._ids[0],
            }  
        #
        if self.option == 'cm':
            contracts_not_found = []
            payments_not_found = []
            valores = []
            i = 0
            for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=None, values_only=True):                  
                # Se busca el pago
                payment_id = self.env['account.payment'].search([('name','ilike',record[3]),('company_id','=',self.company_id.id)])
                #
                if payment_id:
                    name = record[5]
                    if name == 'PAPELERIA':
                        name = 'PAPELERIA PABS'
                    if name == 'FIDEICOMISO':
                        name = 'FIDEICOMISO PABS' 
                    agent_id = self.env['hr.employee'].search([('barcode','=',record[6]),('company_id','=',self.company_id.id)], limit = 1)                               
                    #
                    jobs = {
                        'ASISTENTE SOCIAL': 810,
                        'RECOMENDADO': 811,
                        'COORDINADOR': 812,
                        'GERENTE DE OFICINA': 813,
                        'FIDEICOMISO': 814,
                        'PAPELERIA': 815,
                        'COBRADOR': 816,
                    }
                    if agent_id:
                        vals = {
                            'payment_id': payment_id.id,                              
                            'job_id': jobs.get(record[5]),
                            'comission_agent_id': agent_id.id,                                                            
                            'commission_paid': record[8],
                            'actual_commission_paid': record[9],
                        }  
                        valores.append(vals)                     
                        # self.env['pabs.comission.output'].create(vals)
                        i += 1
                        _logger.info("Creando comisión %s "%(i))
                else:
                    if record[3] not in payments_not_found:
                        payments_not_found.append(record[3])
                        _logger.info("Pago no Encontrado: %s"%record[3])      
            
            # self.env['pabs.comission.output'].create(valores)
            self.data = str(valores)
            self.info = str(contracts_not_found) + "\n" + str(payments_not_found) 
            #
            return {
                    'name':"Importar XLS",
                    'view_type': 'form',
                    'view_mode': 'form',
                    'view_id': False,
                    'res_model': 'import.xls.wizard',
                    'domain': [],
                    'type': 'ir.actions.act_window',
                    'target': 'new',
                    'res_id': self._ids[0],
            }  