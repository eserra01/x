# -*- encoding: utf-8 -*-
from odoo import models, fields,_
import openpyxl
import base64
from io import BytesIO
from odoo.exceptions import UserError
from odoo.exceptions import ValidationError
import logging

_logger = logging.getLogger(__name__)

class PabsEleanorMoveImportImportXLSWizard(models.Model):
    _name = 'pabs.eleanor.move.import.xls.wizard'
    
    file = fields.Binary(string="Archivo", required=True)
    file_name = fields.Char(string="Archivo")
    period_type = fields.Selection([('weekly','Semanal'),('biweekly','Quincenal')], string="Tipo", required=True)         
    info = fields.Char(string="Resultados",default="", readonly=True)         

    def import_file(self):
        id_compania = self.env.company.id
        _logger.info("Comienza carga de plantilla de Eleanor compañia {}".format(id_compania))

        mov_obj = self.env['pabs.eleanor.move']
        conc_obj = self.env['pabs.eleanor.concept']

        ### Validar que exista un periodo abierto
        period_id = self.env['pabs.eleanor.move'].get_period(period_type=self.period_type)
        if not period_id:
            raise ValidationError("No se encontró un periodo abierto")
        else:
            period_id = period_id.id

        ### Leer archivo
        wb = openpyxl.load_workbook( 
        filename=BytesIO(base64.b64decode(self.file)), read_only=True)
        ws = wb.active
                  
        headers = ws.iter_rows(min_row=1, max_row=1, min_col=None,max_col=None, values_only=True)
        records = ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=None, values_only=True)

        ### Validar los conceptos del archivo VS los conceptos de Odoo y generar una lista. (los conceptos comienzan en la columna 5)
        conceptos_odoo = []
        for fila in headers:
            for celda in fila[4:9999]:
                #Dejar de buscar en columnas en la primera columna en blanco
                if celda == None:
                    break

                concepto = conc_obj.search([
                        ('company_id', '=', id_compania),
                        ('allow_load', '=', True),
                        ('name','=', celda)
                    ])
                
                if not concepto:
                    raise ValidationError("No se encuentra el concepto: {}".format(celda))
                elif len(concepto) > 1:
                    raise ValidationError("El concepto se encuentra en más de una columna: {}".format(celda))
                else:
                    conceptos_odoo.append(concepto)

        ### Obtener permisos de acceso del usuario ###
        ids_departamentos = []
        ids_oficinas = []
        all_employees_allowed = self.env.user.all_employees

        if not all_employees_allowed:
            accesos = self.env['pabs.eleanor.user.access'].search([
                ('company_id', '=', id_compania),
                ('user_id', '=', self.env.user.id)
            ])

            ids_departamentos = accesos.mapped('department_id').ids
            ids_oficinas = accesos.mapped('warehouse_id').ids

        ### Obtener cantidad de registros
        cantidad_registros = 0
        for fila in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=None, values_only=True):
            ### Si se encuentra una fila donde el código esté en blanco asumir que es el final del archivo
            if fila[0] == None:
                break
            else:
                cantidad_registros = cantidad_registros + 1

        ### Obtener movimientos del periodo
        movimientos = []
        for res in mov_obj.search([('period_id','=', period_id), ('company_id','=',id_compania)]):
            movimientos.append({
                'id': res.id,
                'concept_id': res.concept_id.id,
                'employee_id': res.employee_id.id
            })

        ### Comenzar a iterar en cada fila. Al terminar se eliminaran los movimientos existentes y se crearán los nuevos.
        movimientos_por_eliminar = []
        movimientos_por_crear = []
        n = 0
        for n, record in enumerate(records, 1):
            ### Si se encuentra una fila donde el código esté en blanco asumir que es el final del archivo
            if record[0] == None:
                break

            _logger.info("Eleanor carga masiva - {} de {}".format(n, cantidad_registros))

            # Se busca el empleado
            employee_id = self.env['hr.employee'].search(
            [
                ('barcode','=',record[0]),
                ('period_type', '=', self.period_type)
            ], limit=1)
            
            if not employee_id:
                raise ValidationError("No se encuentra el empleado con el código {} - {} o su tipo de periodo no es el correcto".format(record[0], record[1]))
            
            if not employee_id.warehouse_id and not employee_id.department_id:
                raise ValidationError("El empleado {}-{} no tiene asignada una oficina o un departamento.".format(employee_id.barcode, employee_id.name))
            
            # Se busca si el usuario tiene acceso al empleado
            if not all_employees_allowed:
                lugar = ""
                access_id = False

                if employee_id.warehouse_id:
                    if employee_id.warehouse_id.id in ids_oficinas:
                        access_id = True
                    else:
                        lugar = "a la oficina {}".format(employee_id.warehouse_id.name)
                elif employee_id.department_id:
                    if employee_id.department_id.id in ids_departamentos:
                        access_id = True
                    else:
                        lugar = "al departamento {}".format(employee_id.department_id.name)
                    
                if not access_id:
                    raise ValidationError("No puedes crear movimientos para el empleado {}-{} porque no tienes acceso {}".format(employee_id.barcode, employee_id.name, lugar))

            i = 4
            for conc in conceptos_odoo:
                
                ### Se busca un registro existente con el periodo, empleado y concepto para agregar a la lista de movimientos a eliminar. Las celdas en blanco tambien eliminan el registro.
                mov_exi = next((x for x in movimientos if x['employee_id'] == employee_id.id and x['concept_id'] == conc.id), 0)

                if mov_exi:
                    movimientos_por_eliminar.append(mov_exi['id'])
                    movimientos.remove(mov_exi) ### Quitar de la lista para aumentar velocidad de búsqueda

                # Se valida que el valor sea un número
                _type = str(type(record[i]))
                if 'int' in _type or 'float' in _type or 'None' in _type:
                    if 'int' in _type or 'float' in _type:

                        ### No importar movimientos en cero
                        if record[i] == 0:
                            continue
                    
                        ### Validar area del empleado (buscar en oficina, si no tiene buscar en departamento)
                        lugar = ""
                        id_area = 0
                        if employee_id.warehouse_id:
                            if employee_id.warehouse_id.pabs_eleanor_area_id:
                                id_area = employee_id.warehouse_id.pabs_eleanor_area_id.id
                            else:
                                lugar = "La oficina {}".format(employee_id.warehouse_id.name)
                        elif employee_id.department_id:
                            if employee_id.department_id.pabs_eleanor_area_id:
                                id_area = employee_id.department_id.pabs_eleanor_area_id.id
                            else:
                                lugar = "El departamento {}".format(employee_id.department_id.name)

                        if not id_area:
                            raise ValidationError("{} del empleado {}-{} no tiene asignada un area.".format(lugar, employee_id.barcode, employee_id.name))
                        
                        ### Validar puesto del empleado
                        if not employee_id.job_id: 
                            raise ValidationError("El empleado {}-{} no tiene asignado un puesto".format(employee_id.barcode, employee_id.name))

                        ### Agregar registro a lista de movimientos por crear
                        movimientos_por_crear.append({  
                            'period_id': period_id,
                            'move_type': conc.concept_type,
                            'concept_id': conc.id,
                            'employee_id': employee_id.id,
                            'area_id': id_area,
                            'warehouse_id': employee_id.warehouse_id.id,
                            'department_id': employee_id.department_id.id,
                            'job_id': employee_id.job_id.id,
                            'amount': record[i],
                            'company_id': id_compania
                        })                        
                else:
                    raise ValidationError("El valor de una celda no es válido: {}. Empleado {}-{}. Concepto {}".format(record[i], employee_id.barcode, employee_id.name, conc.name))
                
                i+= 1

        ### Eliminar movimientos existentes
        if movimientos_por_eliminar:
            _logger.info('Comienza eliminación de movimientos')
            rs_movimientos = mov_obj.browse(movimientos_por_eliminar)
            rs_movimientos.with_context({'byfile': True}).unlink()

        ### Crear movimentos
        if movimientos_por_crear:
            _logger.info('Comienza creación de movimientos')
            mov_obj.with_context({'masive': True}).create(movimientos_por_crear)

        # Se devuelven los resultados de los registros insertados 
        self.info = "Se importaron los movimientos de {} empleados.".format(n)
        return {
                'name':"Resultados de importación",
                'view_type': 'form',
                'view_mode': 'form',
                'view_id': False,
                'res_model': 'pabs.eleanor.move.import.xls.wizard',
                'domain': [],
                'type': 'ir.actions.act_window',
                'target': 'new',
                'res_id': self._ids[0],
        }